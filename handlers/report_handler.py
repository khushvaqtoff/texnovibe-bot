"""
Hisobot, Ombor nazorati va Excel eksport handleri
Yangilik:
  - Savdolarda va to'lovlarda Ish Joyi ustuni
  - Har bir jadval tagida umumiy hisob qatori
"""

from telegram import Update
from telegram.ext import ContextTypes
from sheets.google_sheets import get_spreadsheet, ensure_worksheets, ws_to_records
from datetime import date, datetime
import io
import os


def format_money(amount) -> str:
    try:
        return f"{int(float(amount)):,}".replace(",", " ")
    except:
        return str(amount)


def safe_float(val):
    try:
        return float(str(val).replace(" ", "").replace(",", "").strip() or 0)
    except:
        return 0


async def cmd_daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Bugungi hisobot yuklanmoqda...")
    try:
        sh           = get_spreadsheet()
        sheets       = ensure_worksheets(sh)
        today_str    = date.today().strftime("%d.%m.%Y")
        all_sales    = ws_to_records(sheets["Savdolar"])
        all_payments = ws_to_records(sheets["Tolovlar"])

        today_sales    = [r for r in all_sales    if r.get("Sana") == today_str]
        today_payments = [r for r in all_payments if str(r.get("To'lov Sanasi", "")) == today_str]

        today_sales_sum   = sum(safe_float(r.get("Jami Summa", 0))              for r in today_sales)
        today_sales_avans = sum(safe_float(r.get("Boshlang'ich To'lov", 0))     for r in today_sales)
        today_pay_sum     = sum(safe_float(r.get("To'lov Summasi", 0))          for r in today_payments)
        total_income      = today_sales_avans + today_pay_sum

        text  = "BUGUNGI HISOBOT\nSana: " + today_str + "\n" + "=" * 30 + "\n\n"
        text += "YANGI SAVDOLAR (" + str(len(today_sales)) + " ta)\n" + "-" * 25 + "\n"

        if today_sales:
            for rec in today_sales:
                work  = rec.get("Ish Joyi", "")
                text += (
                    "+ " + rec.get("FIO", "") + "\n"
                    "  " + rec.get("Tovar", "") + "\n"
                    "  Jami: " + format_money(rec.get("Jami Summa", 0)) +
                    " so'm | Avans: " + format_money(rec.get("Boshlang'ich To'lov", 0)) + " so'm\n"
                    "  " + rec.get("To'lov Turi", "") +
                    (" | Ish joyi: " + work if work else "") + "\n\n"
                )
        else:
            text += "Bugun yangi savdo yo'q\n\n"

        text += "TUSHGAN TO'LOVLAR (" + str(len(today_payments)) + " ta)\n" + "-" * 25 + "\n"
        if today_payments:
            for rec in today_payments:
                text += (
                    "+ " + rec.get("FIO", "") + "\n"
                    "  To'landi: " + format_money(rec.get("To'lov Summasi", 0)) + " so'm\n"
                    "  Qoldiq: "   + format_money(rec.get("Qoldiq", 0))          + " so'm\n\n"
                )
        else:
            text += "Bugun to'lov tushgani yo'q\n\n"

        text += (
            "=" * 30 + "\nBUGUNGI NATIJA\n"
            "Yangi savdolar: " + format_money(today_sales_sum)   + " so'm\n"
            "Avans tushdi: "   + format_money(today_sales_avans) + " so'm\n"
            "To'lovlar tushdi: " + format_money(today_pay_sum)   + " so'm\n"
            "Jami kassa: "     + format_money(total_income)       + " so'm\n"
        )

        for i in range(0, len(text), 4000):
            await update.message.reply_text(text[i:i+4000])

    except Exception as e:
        await update.message.reply_text("Xatolik: " + str(e))


async def cmd_warehouse(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Ombor ma'lumotlari yuklanmoqda...")
    try:
        sh           = get_spreadsheet()
        sheets       = ensure_worksheets(sh)
        today_str    = date.today().strftime("%d.%m.%Y")
        all_sales    = ws_to_records(sheets["Savdolar"])
        all_payments = ws_to_records(sheets["Tolovlar"])

        faol      = [r for r in all_sales if r.get("Holat") == "Faol"]
        yopilgan  = [r for r in all_sales if r.get("Holat") == "Yopildi"]
        bekor     = [r for r in all_sales if r.get("Holat") == "Bekor qilindi"]
        hisob     = [r for r in all_sales if r.get("Holat") in ("Faol","Yopildi")]

        jami_nasiya  = sum(safe_float(r.get("Jami Summa",0))           for r in hisob)
        jami_qoldiq  = sum(safe_float(r.get("Qoldiq",0))               for r in faol)
        jami_tolangan= sum(safe_float(r.get("To'langan Summa",0))      for r in hisob)
        jami_avans   = sum(safe_float(r.get("Boshlang'ich To'lov",0))  for r in hisob)

        bugungi     = [r for r in all_payments if str(r.get("To'lov Sanasi","")) == today_str]
        bugungi_sum = sum(safe_float(r.get("To'lov Summasi",0)) for r in bugungi)

        tovar_stats = {}
        for rec in all_sales:
            if rec.get("Holat") == "Bekor qilindi":
                continue
            t = rec.get("Tovar","Noma'lum")
            if t not in tovar_stats:
                tovar_stats[t] = {"faol":0,"yopilgan":0}
            if rec.get("Holat") == "Faol":
                tovar_stats[t]["faol"] += 1
            elif rec.get("Holat") == "Yopildi":
                tovar_stats[t]["yopilgan"] += 1

        text  = "OMBOR NAZORATI\nSana: " + today_str + "\n" + "=" * 30 + "\n\n"
        text += "UMUMIY HOLAT\n" + "-"*25 + "\n"
        text += "Jami sotilgan: " + str(len(all_sales)) + " ta\n"
        text += "Faol nasiya: "   + str(len(faol))      + " ta\n"
        text += "Yopilgan: "      + str(len(yopilgan))  + " ta\n"
        text += "Bekor: "         + str(len(bekor))     + " ta\n\n"
        text += "MOLIYAVIY HOLAT\n" + "-"*25 + "\n"
        text += "Jami nasiya: "   + format_money(jami_nasiya)   + " so'm\n"
        text += "Jami avans: "    + format_money(jami_avans)    + " so'm\n"
        text += "Jami to'langan: "+ format_money(jami_tolangan) + " so'm\n"
        text += "Qolgan qarz: "   + format_money(jami_qoldiq)   + " so'm\n\n"
        text += "BUGUNGI TO'LOVLAR (" + str(len(bugungi)) + " ta)\n" + "-"*25 + "\n"
        text += "Bugun tushdi: "  + format_money(bugungi_sum)   + " so'm\n"
        for rec in bugungi:
            text += "  + " + rec.get("FIO","") + ": " + format_money(rec.get("To'lov Summasi",0)) + " so'm\n"
        text += "\nTOVAR BO'YICHA\n" + "-"*25 + "\n"
        for t, s in sorted(tovar_stats.items(), key=lambda x: x[1]["faol"]+x[1]["yopilgan"], reverse=True)[:10]:
            text += t + ": Faol " + str(s["faol"]) + " | Yopilgan " + str(s["yopilgan"]) + "\n"

        for i in range(0, len(text), 4000):
            await update.message.reply_text(text[i:i+4000])

    except Exception as e:
        await update.message.reply_text("Xatolik: " + str(e))


async def cmd_excel_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Excel fayl tayyorlanmoqda...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        sh           = get_spreadsheet()
        sheets       = ensure_worksheets(sh)
        today_str    = date.today().strftime("%d.%m.%Y")
        all_sales    = ws_to_records(sheets["Savdolar"])
        all_payments = ws_to_records(sheets["Tolovlar"])
        all_clients  = ws_to_records(sheets["Mijozlar"])

        wb = Workbook()

        BLUE_FILL   = PatternFill("solid", fgColor="2E86AB")
        RED_FILL    = PatternFill("solid", fgColor="FF6B6B")
        YELLOW_FILL = PatternFill("solid", fgColor="FFE66D")
        GREEN_FILL  = PatternFill("solid", fgColor="A8E6CF")
        TOTAL_FILL  = PatternFill("solid", fgColor="D5E8D4")
        WHITE_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        BOLD        = Font(bold=True, name="Arial", size=10)
        NORMAL      = Font(name="Arial", size=10)
        CENTER      = Alignment(horizontal="center", vertical="center")
        thin        = Side(style="thin", color="CCCCCC")
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(cell, fill=None):
            cell.font      = WHITE_FONT
            cell.fill      = fill or BLUE_FILL
            cell.alignment = CENTER
            cell.border    = BORDER

        def style_cell(cell, bold=False, fill=None):
            cell.font      = BOLD if bold else NORMAL
            cell.alignment = Alignment(vertical="center")
            cell.border    = BORDER
            if fill:
                cell.fill = fill

        def style_total(cell):
            cell.font      = BOLD
            cell.fill      = TOTAL_FILL
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")

        def set_widths(ws, widths):
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

        today_date   = date.today()
        faol_sales   = [r for r in all_sales if r.get("Holat") == "Faol"]
        today_sales  = [r for r in all_sales if r.get("Sana") == today_str]
        today_pays   = [r for r in all_payments if str(r.get("To'lov Sanasi","")) == today_str]

        # ══════════════════════════════════════════
        # 1. QARZDORLAR
        # ══════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Qarzdorlar"
        h1 = ["#","FIO","Telefon","Tovar","Ish Joyi","Jami Summa",
              "Qoldiq","To'lov Turi","Keyingi To'lov","Kechikish (kun)","Agent"]
        for j, h in enumerate(h1, 1):
            style_header(ws1.cell(row=1, column=j, value=h))

        row = 2
        jami_q = jami_j = 0
        for idx, rec in enumerate(faol_sales, 1):
            try:
                nps    = str(rec.get("Keyingi To'lov Sanasi",""))
                npd    = datetime.strptime(nps, "%d.%m.%Y").date() if nps else None
                delay  = max(0, (today_date - npd).days) if npd else 0
            except:
                delay = 0
            fill  = RED_FILL if delay >= 3 else (YELLOW_FILL if delay >= 1 else GREEN_FILL)
            jami  = safe_float(rec.get("Jami Summa", 0))
            qold  = safe_float(rec.get("Qoldiq", 0))
            jami_j += jami
            jami_q += qold
            vals  = [idx, rec.get("FIO",""), rec.get("Telefon",""), rec.get("Tovar",""),
                     rec.get("Ish Joyi",""), jami, qold,
                     rec.get("To'lov Turi",""), rec.get("Keyingi To'lov Sanasi",""),
                     delay, rec.get("Agent","")]
            for j, val in enumerate(vals, 1):
                style_cell(ws1.cell(row=row, column=j, value=val), bold=(j==2), fill=fill)
            row += 1

        # Jami qator
        ws1.cell(row=row, column=1, value="JAMI")
        ws1.cell(row=row, column=6, value=jami_j)
        ws1.cell(row=row, column=7, value=jami_q)
        for j in range(1, 12):
            style_total(ws1.cell(row=row, column=j))

        set_widths(ws1, {"A":5,"B":22,"C":16,"D":18,"E":18,
                         "F":14,"G":14,"H":12,"I":14,"J":14,"K":14})

        # ══════════════════════════════════════════
        # 2. MIJOZLAR
        # ══════════════════════════════════════════
        ws2 = wb.create_sheet("Mijozlar")
        h2  = ["#","FIO","Telefon","Jami Savdolar","Status","Kredit Bali","Ish Joyi","Sana"]
        for j, h in enumerate(h2, 1):
            style_header(ws2.cell(row=1, column=j, value=h))
        for i, rec in enumerate(all_clients, 1):
            vals = [i, rec.get("FIO",""), rec.get("Telefon",""),
                    safe_float(rec.get("Jami Savdolar",0)), rec.get("Status","Bronze"),
                    safe_float(rec.get("Kredit Bali",0)), rec.get("Ish Joyi",""),
                    rec.get("Ro'yxatga Olingan Sana","")]
            for j, val in enumerate(vals, 1):
                style_cell(ws2.cell(row=i+1, column=j, value=val), bold=(j==2))

        # Jami qator
        tr = len(all_clients) + 2
        ws2.cell(row=tr, column=1, value="JAMI")
        ws2.cell(row=tr, column=4, value=sum(safe_float(r.get("Jami Savdolar",0)) for r in all_clients))
        for j in range(1, 9):
            style_total(ws2.cell(row=tr, column=j))
        set_widths(ws2, {"A":5,"B":22,"C":16,"D":14,"E":12,"F":12,"G":18,"H":18})

        # ══════════════════════════════════════════
        # 3. BUGUNGI HISOBOT
        # ══════════════════════════════════════════
        ws3 = wb.create_sheet("Bugungi Hisobot")
        ws3.merge_cells("A1:H1")
        c = ws3.cell(row=1, column=1, value="BUGUNGI HISOBOT -- " + today_str)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        c.fill = BLUE_FILL
        c.alignment = CENTER

        # — Yangi savdolar —
        ws3.cell(row=2, column=1, value="YANGI SAVDOLAR").font = BOLD
        h3a = ["#","FIO","Telefon","Ish Joyi","Tovar","Jami Summa","Avans","To'lov Turi"]
        for j, h in enumerate(h3a, 1):
            c = ws3.cell(row=3, column=j, value=h)
            c.font = WHITE_FONT; c.fill = GREEN_FILL
            c.alignment = CENTER; c.border = BORDER

        sr = 4
        jami_s = avans_s = 0
        for i, rec in enumerate(today_sales, 1):
            jami  = safe_float(rec.get("Jami Summa",0))
            avans = safe_float(rec.get("Boshlang'ich To'lov",0))
            jami_s  += jami
            avans_s += avans
            vals = [i, rec.get("FIO",""), rec.get("Telefon",""), rec.get("Ish Joyi",""),
                    rec.get("Tovar",""), jami, avans, rec.get("To'lov Turi","")]
            for j, val in enumerate(vals, 1):
                style_cell(ws3.cell(row=sr, column=j, value=val))
            sr += 1

        # Savdolar jami
        ws3.cell(row=sr, column=1, value="JAMI")
        ws3.cell(row=sr, column=6, value=jami_s)
        ws3.cell(row=sr, column=7, value=avans_s)
        for j in range(1, 9):
            style_total(ws3.cell(row=sr, column=j))
        sr += 2

        # — To'lovlar —
        ws3.cell(row=sr, column=1, value="TUSHGAN TO'LOVLAR").font = BOLD
        sr += 1
        h3b = ["#","FIO","Telefon","Ish Joyi","Savdo ID","To'lov Summasi","Qoldiq","Sana"]
        for j, h in enumerate(h3b, 1):
            c = ws3.cell(row=sr, column=j, value=h)
            c.font = WHITE_FONT; c.fill = YELLOW_FILL
            c.alignment = CENTER; c.border = BORDER
        sr += 1

        pay_sum = qold_sum = 0
        for i, rec in enumerate(today_pays, 1):
            # Ish joyini savdolardan topish
            phone_clean = str(rec.get("Telefon","")).replace("+","").replace(" ","").replace("-","")
            ish_joyi = ""
            for s in all_sales:
                sp = str(s.get("Telefon","")).replace("+","").replace(" ","").replace("-","")
                if sp == phone_clean and str(s.get("ID","")) == str(rec.get("Savdo ID","")):
                    ish_joyi = s.get("Ish Joyi","")
                    break

            p   = safe_float(rec.get("To'lov Summasi",0))
            q   = safe_float(rec.get("Qoldiq",0))
            pay_sum  += p
            qold_sum += q
            vals = [i, rec.get("FIO",""), rec.get("Telefon",""), ish_joyi,
                    rec.get("Savdo ID",""), p, q, rec.get("To'lov Sanasi","")]
            for j, val in enumerate(vals, 1):
                style_cell(ws3.cell(row=sr, column=j, value=val))
            sr += 1

        # To'lovlar jami
        ws3.cell(row=sr, column=1, value="JAMI")
        ws3.cell(row=sr, column=6, value=pay_sum)
        ws3.cell(row=sr, column=7, value=qold_sum)
        for j in range(1, 9):
            style_total(ws3.cell(row=sr, column=j))
        sr += 2

        # Umumiy kassa
        summary = [
            ("Yangi savdolar:", len(today_sales)),
            ("To'lovlar:", len(today_pays)),
            ("Avans tushdi:", avans_s),
            ("To'lovlar tushdi:", pay_sum),
            ("JAMI KASSA:", avans_s + pay_sum),
        ]
        for label, val in summary:
            ws3.cell(row=sr, column=1, value=label).font = BOLD
            c = ws3.cell(row=sr, column=2, value=val)
            c.font = BOLD
            if label == "JAMI KASSA:":
                ws3.cell(row=sr, column=1).fill = TOTAL_FILL
                c.fill = TOTAL_FILL
            sr += 1

        set_widths(ws3, {"A":5,"B":22,"C":16,"D":18,"E":14,"F":16,"G":14,"H":14})

        # ══════════════════════════════════════════
        # Saqlash va yuborish
        # ══════════════════════════════════════════
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "texnovibe_" + today_str.replace(".", "_") + ".xlsx"
        await update.message.reply_document(
            document=buf,
            filename=filename,
            caption=(
                "Excel hisobot -- " + today_str + "\n"
                "Qarzdorlar | Mijozlar | Bugungi Hisobot\n"
                "Savdolarda va to'lovlarda Ish Joyi qo'shildi"
            )
        )

    except Exception as e:
        await update.message.reply_text("Xatolik: " + str(e))
