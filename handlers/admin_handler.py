"""
Admin handlerlari
/eksport — Excel fayl eksport
/backup — Zaxira nusxa
"""

import os
import io
from datetime import date
from telegram import Update
from telegram.ext import ContextTypes
from sheets.google_sheets import get_spreadsheet, ensure_worksheets

ADMIN_CHAT_ID = int(os.getenv("ADMIN_CHAT_ID", "0"))


async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/eksport — Barcha ma'lumotlarni Excel formatida yuboradi"""
    if update.effective_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ Bu buyruq faqat admin uchun.")
        return

    await update.message.reply_text("⏳ Excel fayl tayyorlanmoqda...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        sh = get_spreadsheet()
        sheets = ensure_worksheets(sh)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Default sheetni o'chirish

        # Har bir listni eksport qilish
        for sheet_name, ws_remote in sheets.items():
            ws_local = wb.create_sheet(title=sheet_name)
            data = ws_remote.get_all_values()

            if not data:
                continue

            # Sarlavha
            header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, cell_value in enumerate(data[0], 1):
                cell = ws_local.cell(row=1, column=col_idx, value=cell_value)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Ma'lumotlar
            for row_idx, row in enumerate(data[1:], 2):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws_local.cell(row=row_idx, column=col_idx, value=cell_value)
                    # Juft satrlar uchun rang
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(
                            start_color="F0F8FF",
                            end_color="F0F8FF",
                            fill_type="solid"
                        )

            # Ustun kengliklarini avtomatik sozlash
            for col in ws_local.columns:
                max_length = max((len(str(cell.value or "")) for cell in col), default=10)
                ws_local.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 40)

        # Faylni xotiraga saqlash
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        filename = f"TexnoVibe_{date.today().strftime('%d-%m-%Y')}.xlsx"

        await update.message.reply_document(
            document=file_buffer,
            filename=filename,
            caption=(
                f"✅ *Excel eksport tayyor!*\n"
                f"📅 Sana: {date.today().strftime('%d.%m.%Y')}\n"
                f"📊 Barcha ma'lumotlar kiritilgan."
            ),
            parse_mode="Markdown"
        )

    except ImportError:
        await update.message.reply_text(
            "❌ `openpyxl` kutubxonasi o'rnatilmagan.\n"
            "Terminal da: `pip install openpyxl`",
            parse_mode="Markdown"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Eksport xatosi: `{str(e)}`", parse_mode="Markdown")


async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/backup — Spreadsheet havolasini yuboradi"""
    if update.effective_user.id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ Bu buyruq faqat admin uchun.")
        return

    spreadsheet_id = os.getenv("SPREADSHEET_ID", "")

    if spreadsheet_id:
        link = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        await update.message.reply_text(
            f"☁️ *Google Sheets Zaxira*\n\n"
            f"📊 [TexnoVibe Nasiya Baza]({link})\n\n"
            f"Ma'lumotlaringiz real vaqtda Google Sheets da saqlanmoqda.\n"
            f"Istalgan vaqt yuqoridagi havoladan ko'ring.",
            parse_mode="Markdown",
            disable_web_page_preview=True
        )
    else:
        await update.message.reply_text(
            "⚠️ SPREADSHEET_ID sozlanmagan.\n"
            "`.env` faylini tekshiring."
        )
